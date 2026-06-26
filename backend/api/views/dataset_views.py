"""
Dataset API Views
==================
CRUD + JSONL/CSV/XLSX upload + sample management for datasets.
"""

import csv
import hashlib
import io
import json
import logging
import os
import uuid

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from evaluation.models import Dataset, DatasetSample
from api.serializers import (
    DatasetSerializer,
    DatasetDetailSerializer,
    DatasetSampleSerializer,
    DatasetSampleUploadSerializer,
)

logger = logging.getLogger(__name__)

# Column name mapping for CSV/XLSX golden dataset format
# convId + userId + content -> input JSON
# expected_output + toolcalling -> expected_output JSON
# expected_meta -> expected_meta JSON
GOLDEN_INPUT_COLS = {"convId", "conv_id", "convid"}  # accepted variants
GOLDEN_USER_COLS = {"userId", "user_id", "userid"}
GOLDEN_CONTENT_COLS = {"content", "query", "question", "input"}
GOLDEN_OUTPUT_COLS = {"expected_output", "expected", "answer", "output"}
GOLDEN_TOOL_COLS = {"toolcalling", "tool_calling", "tool", "tool_call"}
GOLDEN_META_COLS = {"expected_meta", "expectedmeta", "meta", "meta_expected"}


def _find_col(header, candidates):
    """Find the index of a column matching any candidate name (case-insensitive)."""
    for i, h in enumerate(header):
        if h.strip() in candidates or h.strip().lower() in {c.lower() for c in candidates}:
            return i
    return None


def _parse_meta_column(value: str) -> dict:
    """Parse expected_meta column value. Accepts JSON string, key:value pairs, or empty."""
    if not value or not value.strip():
        return {}
    raw = value.strip()
    # Try direct JSON parse first
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, dict):
            return parsed
        return {}
    except (json.JSONDecodeError, TypeError):
        pass
    # Fallback: wrap bare key:value pair(s) in braces, e.g. "agentId":"xxx" -> {"agentId":"xxx"}
    try:
        wrapped = "{" + raw + "}"
        parsed = json.loads(wrapped)
        if isinstance(parsed, dict):
            return parsed
    except (json.JSONDecodeError, TypeError):
        pass
    return {}


def _parse_spreadsheet_rows(rows_with_header):
    """
    Parse spreadsheet rows into sample dicts.
    First element of rows_with_header is the header row.
    Returns (samples: list[dict], errors: list[dict]).
    """
    if not rows_with_header:
        return [], [{"line": 0, "error": "File is empty"}]

    header = [str(c).strip() if c else "" for c in rows_with_header[0]]
    errors = []
    samples = []

    col_conv = _find_col(header, GOLDEN_INPUT_COLS)
    col_user = _find_col(header, GOLDEN_USER_COLS)
    col_content = _find_col(header, GOLDEN_CONTENT_COLS)
    col_output = _find_col(header, GOLDEN_OUTPUT_COLS)
    col_tool = _find_col(header, GOLDEN_TOOL_COLS)
    col_meta = _find_col(header, GOLDEN_META_COLS)

    # Fallback: if no content column found, try column index 2 (typical layout)
    if col_content is None and len(header) >= 3:
        col_content = 2

    for line_num, row in enumerate(rows_with_header[1:], 2):
        cells = [str(c).strip() if c else "" for c in row]
        if not any(cells):
            continue  # skip empty rows

        try:
            # Build input JSON
            input_obj = {}
            if col_conv is not None and col_conv < len(cells) and cells[col_conv]:
                input_obj["convId"] = cells[col_conv]
            if col_user is not None and col_user < len(cells) and cells[col_user]:
                input_obj["userId"] = cells[col_user]
            if col_content is not None and col_content < len(cells) and cells[col_content]:
                input_obj["content"] = cells[col_content]

            if not input_obj.get("content"):
                errors.append({"line": line_num, "error": "Missing content/query column"})
                continue

            # Build expected_output: store as plain text if no toolcalling, otherwise wrap in JSON
            output_text = cells[col_output] if (col_output is not None and col_output < len(cells) and cells[col_output]) else ""
            if col_tool is not None and col_tool < len(cells) and cells[col_tool]:
                output_obj = {}
                if output_text:
                    output_obj["expected_output"] = output_text
                output_obj["toolcalling"] = cells[col_tool]
                output_text = json.dumps(output_obj, ensure_ascii=False)

            # Derive sample_id from convId + content hash for uniqueness
            conv_id = input_obj.get("convId", "")
            content_hash = hashlib.md5(input_obj["content"].encode()).hexdigest()[:8]
            sample_id = f"{conv_id}_{content_hash}" if conv_id else f"sample_{uuid.uuid4().hex[:8]}"
            # Sanitize sample_id (only alphanumerics, underscores, hyphens)
            sample_id = "".join(c if c.isalnum() or c in "_-" else "_" for c in sample_id)[:50]

            samples.append({
                "sample_id": sample_id,
                "input": json.dumps(input_obj, ensure_ascii=False),
                "expected_output": output_text,
                "expected_meta": _parse_meta_column(
                    cells[col_meta] if col_meta is not None and col_meta < len(cells) else ""
                ),
                "context": [],
                "tags": [input_obj.get("toolcalling", "")] if input_obj.get("toolcalling") else [],
            })
        except Exception as e:
            errors.append({"line": line_num, "error": str(e)})

    return samples, errors


def _parse_csv_file(content: str):
    """Parse CSV file with header row. Returns (samples, errors)."""
    reader = csv.reader(io.StringIO(content))
    rows = list(reader)
    return _parse_spreadsheet_rows(rows)


def _parse_xlsx_file(file_bytes: bytes):
    """Parse XLSX file using openpyxl. Returns (samples, errors)."""
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    # Use the first sheet
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return _parse_spreadsheet_rows(rows)


class DatasetViewSet(viewsets.ModelViewSet):
    """
    CRUD operations for datasets.

    list:      GET    /api/v1/datasets/
    create:    POST   /api/v1/datasets/
    read:      GET    /api/v1/datasets/{id}/
    update:    PUT    /api/v1/datasets/{id}/
    delete:    DELETE /api/v1/datasets/{id}/
    upload:    POST   /api/v1/datasets/{id}/upload/
    samples:   GET    /api/v1/datasets/{id}/samples/
    add_sample:POST   /api/v1/datasets/{id}/add_sample/
    """
    queryset = Dataset.objects.all().order_by("-created_at")
    serializer_class = DatasetSerializer
    lookup_field = "id"

    def get_serializer_class(self):
        if self.action == "retrieve":
            return DatasetDetailSerializer
        return DatasetSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        status_filter = self.request.query_params.get("status")
        if status_filter:
            qs = qs.filter(status=status_filter)
        search = self.request.query_params.get("search")
        if search:
            qs = qs.filter(name__icontains=search)
        data_type = self.request.query_params.get("data_type")
        if data_type:
            qs = qs.filter(data_type=data_type)
        return qs

    def perform_create(self, serializer):
        dataset = serializer.save()
        dataset.update_sample_count()

    @action(detail=True, methods=["post"], url_path="upload",
            parser_classes=[MultiPartParser, FormParser])
    def upload_file(self, request, id=None):
        """
        Upload dataset file to add samples. Supports JSONL, CSV, and XLSX.

        POST /api/v1/datasets/{id}/upload/
        Content-Type: multipart/form-data
        Body: file=<.jsonl | .csv | .xlsx>, auto_sample_id=true/false

        CSV/XLSX column mapping:
          - convId + userId + content  -> input JSON
          - expected_output + toolcalling -> expected_output JSON
        """
        dataset = self.get_object()

        serializer = DatasetSampleUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        uploaded_file = serializer.validated_data["file"]
        auto_sample_id = serializer.validated_data.get("auto_sample_id", True)
        filename = uploaded_file.name or ""
        ext = os.path.splitext(filename)[1].lower()

        created = 0
        skipped = 0
        errors = []

        try:
            if ext == ".jsonl" or ext == ".json":
                samples, parse_errors = self._parse_jsonl_file(uploaded_file, auto_sample_id)
                errors.extend(parse_errors)
            elif ext == ".csv":
                content = uploaded_file.read().decode("utf-8-sig")
                samples, parse_errors = _parse_csv_file(content)
                errors.extend(parse_errors)
            elif ext in (".xlsx", ".xls"):
                file_bytes = uploaded_file.read()
                samples, parse_errors = _parse_xlsx_file(file_bytes)
                errors.extend(parse_errors)
            else:
                return Response(
                    {"error": f"Unsupported file format: '{ext}'. Use .jsonl, .csv, or .xlsx"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Bulk insert samples
            for sample_data in samples:
                sample_id = sample_data.get("sample_id")
                if not sample_id and auto_sample_id:
                    sample_id = f"sample_{uuid.uuid4().hex[:8]}"
                if not sample_id:
                    continue

                try:
                    DatasetSample.objects.update_or_create(
                        dataset=dataset,
                        sample_id=sample_id,
                        defaults={
                            "input": sample_data.get("input", ""),
                            "expected_output": sample_data.get("expected_output", ""),
                            "expected_meta": sample_data.get("expected_meta", {}),
                            "context": sample_data.get("context", []),
                            "retrieval_context": sample_data.get("retrieval_context", []),
                            "additional_tools_output": sample_data.get("additional_tools_output", {}),
                            "tags": sample_data.get("tags", []),
                            "notes": sample_data.get("notes", ""),
                        },
                    )
                    created += 1
                except Exception as e:
                    errors.append({"sample_id": sample_id, "error": str(e)})
                    skipped += 1

        except Exception as e:
            logger.exception("Upload failed for file: %s", filename)
            return Response(
                {"error": f"Failed to process file: {e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        dataset.update_sample_count()

        return Response({
            "uploaded": created,
            "skipped": skipped,
            "errors": errors[:20],
            "total_samples": dataset.sample_count,
        })

    @staticmethod
    def _parse_jsonl_file(uploaded_file, auto_sample_id):
        """Parse JSONL file. Returns (samples, errors)."""
        content = uploaded_file.read().decode("utf-8")
        lines = [line.strip() for line in content.splitlines() if line.strip()]
        samples = []
        errors = []
        for line_num, line in enumerate(lines, 1):
            try:
                data = json.loads(line)
                sample_id = data.get("sample_id") or (
                    f"sample_{uuid.uuid4().hex[:8]}" if auto_sample_id else None
                )
                if not sample_id:
                    errors.append({"line": line_num, "error": "Missing sample_id"})
                    continue
                samples.append({
                    "sample_id": sample_id,
                    "input": data.get("input", ""),
                    "expected_output": data.get("expected_output", ""),
                    "expected_meta": data.get("expected_meta", {}),
                    "context": data.get("context", []),
                    "retrieval_context": data.get("retrieval_context", []),
                    "additional_tools_output": data.get("additional_tools_output", {}),
                    "tags": data.get("tags", []),
                    "notes": data.get("notes", ""),
                })
            except json.JSONDecodeError as e:
                errors.append({"line": line_num, "error": f"Invalid JSON: {e}"})
        return samples, errors

    @action(detail=True, methods=["get"], url_path="samples")
    def list_samples(self, request, id=None):
        """
        List samples in the dataset with pagination.

        GET /api/v1/datasets/{id}/samples/?page=1&page_size=20
        """
        dataset = self.get_object()
        samples = dataset.samples.all().order_by("created_at")

        # Pagination
        page = int(request.query_params.get("page", 1))
        page_size = int(request.query_params.get("page_size", 20))
        start = (page - 1) * page_size
        end = start + page_size
        page_samples = samples[start:end]

        serializer = DatasetSampleSerializer(page_samples, many=True)
        return Response({
            "count": samples.count(),
            "page": page,
            "page_size": page_size,
            "results": serializer.data,
        })

    @action(detail=True, methods=["post"], url_path="add_sample")
    def add_sample(self, request, id=None):
        """
        Add a single sample to the dataset.

        POST /api/v1/datasets/{id}/add_sample/
        Body: {"input": "...", "expected_output": "...", "sample_id": "..."}
        """
        import json as _json

        dataset = self.get_object()

        # Pre-process data: normalize fields before serializer validation
        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)

        # Auto-generate sample_id if missing or blank
        if not data.get("sample_id"):
            data["sample_id"] = f"sample_{uuid.uuid4().hex[:8]}"

        # Convert dict/list inputs to JSON strings (frontend sends parsed objects)
        for field in ("input", "expected_output"):
            val = data.get(field)
            if isinstance(val, (dict, list)):
                data[field] = _json.dumps(val, ensure_ascii=False)
            elif val is None:
                data[field] = ""

        # Handle context: accept null, string, or list
        ctx = data.get("context")
        if ctx is None:
            data["context"] = []
        elif isinstance(ctx, str):
            try:
                data["context"] = _json.loads(ctx)
            except (_json.JSONDecodeError, TypeError):
                data["context"] = [ctx]

        serializer = DatasetSampleSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save(dataset=dataset)
        dataset.update_sample_count()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["delete"], url_path="delete_sample/(?P<sample_id>[^/.]+)")
    def delete_sample(self, request, id=None, sample_id=None):
        """
        Delete a single sample from the dataset.

        DELETE /api/v1/datasets/{id}/delete_sample/{sample_id}/
        """
        dataset = self.get_object()
        try:
            sample = dataset.samples.get(sample_id=sample_id)
            sample.delete()
            dataset.update_sample_count()
            return Response({"deleted": sample_id}, status=status.HTTP_200_OK)
        except DatasetSample.DoesNotExist:
            return Response(
                {"error": f"Sample '{sample_id}' not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

    @action(detail=True, methods=["put", "patch", "delete"],
            url_path="samples/(?P<sample_pk>[^/.]+)")
    def sample_detail(self, request, id=None, sample_pk=None):
        """
        Update or delete a single sample by its UUID primary key.

        PUT    /api/v1/datasets/{id}/samples/{sample_pk}/
        PATCH  /api/v1/datasets/{id}/samples/{sample_pk}/
        DELETE /api/v1/datasets/{id}/samples/{sample_pk}/
        """
        dataset = self.get_object()
        try:
            sample = dataset.samples.get(id=sample_pk)
        except DatasetSample.DoesNotExist:
            return Response(
                {"error": f"Sample '{sample_pk}' not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        if request.method == "DELETE":
            sample.delete()
            dataset.update_sample_count()
            return Response({"deleted": sample_pk}, status=status.HTTP_200_OK)

        # PUT or PATCH — update sample
        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)

        # Convert dict/list inputs to JSON strings
        for field in ("input", "expected_output"):
            val = data.get(field)
            if isinstance(val, (dict, list)):
                data[field] = json.dumps(val, ensure_ascii=False)
            elif val is None:
                data[field] = ""

        # Handle context: accept null, string, or list
        ctx = data.get("context")
        if ctx is None:
            data["context"] = []
        elif isinstance(ctx, str):
            try:
                data["context"] = json.loads(ctx)
            except (json.JSONDecodeError, TypeError):
                data["context"] = [ctx]

        partial = request.method in ("PATCH", "PUT")  # Always partial for sample updates
        serializer = DatasetSampleSerializer(sample, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="create_version")
    def create_version(self, request, id=None):
        """
        Create a new version of the dataset (copies current samples).

        POST /api/v1/datasets/{id}/version/
        Body: {"version": "2.0", "copy_samples": true}
        """
        dataset = self.get_object()
        new_version = request.data.get("version", "2.0")
        copy_samples = request.data.get("copy_samples", True)

        new_dataset = Dataset.objects.create(
            name=dataset.name,
            data_type=dataset.data_type,
            version=new_version,
            tags=dataset.tags,
            description=dataset.description,
            status="draft",
        )

        if copy_samples:
            samples = dataset.samples.all()
            new_samples = []
            for s in samples:
                new_samples.append(DatasetSample(
                    dataset=new_dataset,
                    sample_id=s.sample_id,
                    input=s.input,
                    expected_output=s.expected_output,
                    context=s.context,
                    retrieval_context=s.retrieval_context,
                    additional_tools_output=s.additional_tools_output,
                    tags=s.tags,
                    notes=s.notes,
                ))
            DatasetSample.objects.bulk_create(new_samples)
            new_dataset.update_sample_count()

        return Response(
            DatasetSerializer(new_dataset).data,
            status=status.HTTP_201_CREATED,
        )
