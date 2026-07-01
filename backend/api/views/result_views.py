"""
Result, Trace & BadCase API Views
===================================
Query evaluation results, traces, and manage badcase feedback.
"""

import logging

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse

from evaluation.models import EvaluationResult, Trace, BadCaseFeedback, DatasetSample
from api.serializers import (
    EvaluationResultSerializer,
    EvaluationResultListSerializer,
    TraceSerializer,
    BadCaseFeedbackSerializer,
    BadCaseFeedbackCreateSerializer,
)

logger = logging.getLogger(__name__)


class EvaluationResultViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only access to evaluation results.

    list:   GET /api/v1/results/
    read:   GET /api/v1/results/{id}/
    """
    queryset = EvaluationResult.objects.all().select_related("task").order_by("-created_at")
    serializer_class = EvaluationResultSerializer
    lookup_field = "id"

    def get_serializer_class(self):
        if self.action == "list":
            return EvaluationResultListSerializer
        return EvaluationResultSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        task_id = self.request.query_params.get("task")
        if task_id:
            qs = qs.filter(task_id=task_id)
        is_badcase = self.request.query_params.get("is_badcase")
        if is_badcase is not None:
            qs = qs.filter(is_badcase=is_badcase.lower() == "true")
        min_score = self.request.query_params.get("min_score")
        if min_score:
            qs = qs.filter(overall_score__gte=float(min_score))
        max_score = self.request.query_params.get("max_score")
        if max_score:
            qs = qs.filter(overall_score__lte=float(max_score))
        # Filter by sample IDs (comma-separated)
        sample_ids = self.request.query_params.get("sample_ids")
        if sample_ids:
            ids = [sid.strip() for sid in sample_ids.split(",") if sid.strip()]
            if ids:
                qs = qs.filter(sample_id__in=ids)
        # Sorting
        ordering = self.request.query_params.get("ordering", "-overall_score")
        if ordering in ("overall_score", "-overall_score", "latency_ms", "-latency_ms"):
            qs = qs.order_by(ordering)
        return qs

    @action(detail=False, methods=["post"], url_path="filter")
    def filter_results(self, request):
        """
        Filter results via POST (accepts sample_ids in body to avoid URL length limits).

        POST /api/v1/results/filter/
        Body: { "task": "...", "page": 1, "page_size": 20, "sample_ids": ["...", ...], ... }
        """
        # Apply filters from POST body
        qs = EvaluationResult.objects.all().select_related("task").order_by("-created_at")
        task_id = request.data.get("task")
        if task_id:
            qs = qs.filter(task_id=task_id)
        is_badcase = request.data.get("is_badcase")
        if is_badcase is not None:
            qs = qs.filter(is_badcase=str(is_badcase).lower() == "true")
        sample_ids = request.data.get("sample_ids")
        if sample_ids and isinstance(sample_ids, list):
            sample_ids = [sid for sid in sample_ids if sid]
            if sample_ids:
                qs = qs.filter(sample_id__in=sample_ids)
        ordering = request.data.get("ordering", "-overall_score")
        if ordering in ("overall_score", "-overall_score", "latency_ms", "-latency_ms"):
            qs = qs.order_by(ordering)
        
        # Paginate
        from api.pagination import FlexiblePageNumberPagination
        paginator = FlexiblePageNumberPagination()
        page = paginator.paginate_queryset(qs, request)
        if page is not None:
            serializer = EvaluationResultListSerializer(page, many=True)
            return paginator.get_paginated_response(serializer.data)
        
        serializer = EvaluationResultListSerializer(qs[:100], many=True)
        return Response({"count": qs.count(), "results": serializer.data})

    @action(detail=False, methods=["get"], url_path="badcases")
    def list_badcases(self, request):
        """
        List all bad cases across tasks (or filter by task).

        GET /api/v1/results/badcases/?task={task_id}
        """
        qs = self.get_queryset().filter(is_badcase=True)
        serializer = EvaluationResultListSerializer(qs[:100], many=True)
        return Response({
            "count": qs.count(),
            "results": serializer.data,
        })

    @action(detail=False, methods=["get"], url_path="export_jsonl")
    def export_badcases(self, request):
        """
        Export bad cases as JSONL for the specified task.

        GET /api/v1/results/export_jsonl/?task={task_id}
        """
        task_id = request.query_params.get("task")
        if not task_id:
            return Response(
                {"error": "task parameter required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        import json
        badcases = EvaluationResult.objects.filter(
            task_id=task_id, is_badcase=True,
        ).order_by("overall_score")

        lines = []
        for bc in badcases:
            lines.append(json.dumps({
                "input": bc.input,
                "expected_output": bc.expected_output,
                "actual_output": bc.actual_output,
                "overall_score": bc.overall_score,
                "metric_results": bc.metric_results,
                "latency_ms": bc.latency_ms,
            }, ensure_ascii=False))

        content = "\n".join(lines)
        response = HttpResponse(content, content_type="application/x-ndjson")
        response["Content-Disposition"] = f'attachment; filename="badcases_{task_id}.jsonl"'
        return response

    @action(detail=False, methods=["get"], url_path="export_excel")
    def export_excel(self, request):
        """
        Export evaluation results as Excel (.xlsx) for the specified task.

        GET /api/v1/results/export_excel/?task={task_id}
        Columns: is_badcase, input, expected_meta, actual_output, metric_results
        """
        import json
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        task_id = request.query_params.get("task")
        if not task_id:
            return Response(
                {"error": "task parameter required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        results = EvaluationResult.objects.filter(
            task_id=task_id,
        ).select_related("task__dataset").order_by("sample_id").distinct("sample_id")

        # Build a map of sample_id -> expected_meta from DatasetSample
        task_obj = results.first().task if results.exists() else None
        expected_meta_map = {}
        if task_obj and task_obj.dataset:
            samples = DatasetSample.objects.filter(
                dataset=task_obj.dataset,
                sample_id__in=[r.sample_id for r in results],
            ).values("sample_id", "expected_meta")
            expected_meta_map = {s["sample_id"]: s["expected_meta"] for s in samples}

        # Build a map of trace_id -> trace_data from Trace
        trace_ids = [r.trace_id for r in results if r.trace_id]
        trace_data_map = {}
        if trace_ids:
            traces = Trace.objects.filter(trace_id__in=trace_ids).values("trace_id", "trace_data")
            trace_data_map = {t["trace_id"]: t["trace_data"] for t in traces}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "评测结果"

        # ── Header style ──
        header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # ── Headers ──
        headers = ["is_badcase", "input", "expected_meta", "actual_output", "metric_results", "trace_data"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # ── Data rows ──
        data_font = Font(name="微软雅黑", size=10)
        data_alignment = Alignment(vertical="top", wrap_text=True)

        for row_idx, result in enumerate(results, 2):
            expected_meta = expected_meta_map.get(result.sample_id, {})

            trace_data = trace_data_map.get(result.trace_id, {})
            row_data = [
                str(result.is_badcase),
                result.input or "",
                json.dumps(expected_meta, ensure_ascii=False) if expected_meta else "",
                result.actual_output or "",
                json.dumps(result.metric_results, ensure_ascii=False) if result.metric_results else "",
                json.dumps(trace_data, ensure_ascii=False) if trace_data else "",
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        # ── Column widths ──
        ws.column_dimensions["A"].width = 14   # is_badcase
        ws.column_dimensions["B"].width = 50   # input
        ws.column_dimensions["C"].width = 40   # expected_meta
        ws.column_dimensions["D"].width = 60   # actual_output
        ws.column_dimensions["E"].width = 50   # metric_results
        ws.column_dimensions["F"].width = 60   # trace_data

        # ── Freeze header row ──
        ws.freeze_panes = "A2"

        # ── Auto-filter ──
        ws.auto_filter.ref = f"A1:F{len(results) + 1}"

        # ── Response ──
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        task_name = task_obj.name if task_obj else task_id
        filename = f"{task_name}_results.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class TraceViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only access to execution traces.

    list:   GET /api/v1/traces/
    read:   GET /api/v1/traces/{id}/
    """
    queryset = Trace.objects.all().select_related("task").order_by("-created_at")
    serializer_class = TraceSerializer
    lookup_field = "trace_id"

    def get_queryset(self):
        qs = super().get_queryset()
        task_id = self.request.query_params.get("task")
        if task_id:
            qs = qs.filter(task_id=task_id)
        trace_id = self.request.query_params.get("trace_id")
        if trace_id:
            qs = qs.filter(trace_id=trace_id)
        return qs


class BadCaseFeedbackViewSet(viewsets.ModelViewSet):
    """
    CRUD for badcase feedback/annotations.

    list:   GET    /api/v1/feedback/
    create: POST   /api/v1/feedback/
    read:   GET    /api/v1/feedback/{id}/
    update: PUT    /api/v1/feedback/{id}/
    delete: DELETE /api/v1/feedback/{id}/
    """
    queryset = BadCaseFeedback.objects.all().select_related("result", "result__task").order_by("-created_at")
    serializer_class = BadCaseFeedbackSerializer
    lookup_field = "id"

    def get_serializer_class(self):
        if self.action == "create":
            return BadCaseFeedbackCreateSerializer
        return BadCaseFeedbackSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        task_id = self.request.query_params.get("task")
        if task_id:
            qs = qs.filter(result__task_id=task_id)
        result_id = self.request.query_params.get("result")
        if result_id:
            qs = qs.filter(result_id=result_id)
        return qs
